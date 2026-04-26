"""
从 state/orders/*.json 生成闲鱼台账 Excel。

用法：
  python3 scripts/update_xianyu_ledger.py
  python3 scripts/update_xianyu_ledger.py --out xianyu/ledger.xlsx
"""

import argparse
import sys
import unicodedata
from pathlib import Path

_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(_ROOT))

from lib.order_store import list_all
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DEFAULT_OUT = Path("xianyu/ledger.xlsx")

# (header, min_width, max_width, align)
COLUMNS = [
    ("出发日",   12, 12, "left"),
    ("行程",     28, 48, "left"),
    ("人",        4,  4, "right"),
    ("买家",     14, 22, "left"),
    ("成交价",   10, 10, "right"),
    ("携程价",   10, 10, "right"),
    ("利润",     10, 10, "right"),
    ("状态",     10, 10, "left"),
    ("PNR",      14, 26, "left"),
    ("备注",     20, 44, "left"),
]

COL_DEPART, COL_ROUTE, COL_PAX, COL_BUYER, COL_SALE, \
COL_CTRIP, COL_PROFIT, COL_STATUS, COL_PNR, COL_NOTE = range(1, 11)

STATUS_FG = {
    "已收货":   "1A7A45",
    "已发回执": "1A7A45",
    "已出票":   "1A7A45",
    "已付款":   "B07A00",
    "已发单":   "B07A00",
    "已报价":   "5C6470",
    "询价":     "5C6470",
    "交易关闭": "9B1C1C",
}

TYPE_LABEL = {"flight": "机票", "hotel": "酒店", "rail": "铁路"}

C_TEXT          = "1F2328"
C_MUTED         = "6E7781"
C_LINK          = "0969DA"
C_HEADER_BG     = "F6F8FA"
C_BORDER        = "E5E7EB"
C_BORDER_STRONG = "D0D8E4"
C_PROFIT_POS    = "1A7A45"
C_PROFIT_NEG    = "C0392B"

MONEY_FMT = '¥#,##0'
DATE_FMT  = 'yyyy-mm-dd'

COST_RATIO      = 0.4
PROFIT_STATUSES = {"已发回执", "已收货"}

HEADER_ROW_H = 28
DATA_ROW_H   = 26
TOTAL_ROW_H  = 28

FONT_CN = "PingFang SC"


def _bottom_border(color=C_BORDER, weight="thin"):
    return Border(bottom=Side(border_style=weight, color=color))


def _top_border(color=C_BORDER_STRONG, weight="medium"):
    return Border(top=Side(border_style=weight, color=color))


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=C_TEXT, size=12, underline=None):
    return Font(name=FONT_CN, bold=bold, color=color, size=size, underline=underline)


def _cjk_width(s) -> float:
    w = 0.0
    for ch in str(s):
        w += 2.0 if unicodedata.east_asian_width(ch) in ("F", "W") else 1.0
    return w


def _trip_route(o: dict) -> str:
    item      = o.get("item", {}) or {}
    trip      = o.get("trip", {}) or {}
    item_type = item.get("type", "")
    type_tag  = TYPE_LABEL.get(item_type, item_type or "")
    prefix    = f"[{type_tag}] " if type_tag else ""

    if item_type == "rail" and trip.get("route"):
        body = trip["route"]
    elif item_type == "hotel":
        body = trip.get("destination") or trip.get("origin") or item.get("summary", "")
    else:
        def _fmt(name: str, code: str) -> str:
            if name and code: return f"{name} {code}"
            return name or code or ""
        left  = _fmt(trip.get("origin", ""),      trip.get("origin_code", ""))
        right = _fmt(trip.get("destination", ""), trip.get("destination_code", ""))
        if left and right:
            body = f"{left} → {right}"
        elif left or right:
            body = left or right
        else:
            body = item.get("summary", "")

    suffix = ""
    if item_type == "flight":
        suffix = " · 往" if trip.get("is_round_trip") else " · 单"

    return f"{prefix}{body}{suffix}"


def order_to_row(o: dict) -> tuple[list, list[str]]:
    """返回 (单元格值列表, 附件相对路径列表)。"""
    pricing = o.get("pricing", {}) or {}
    buyer   = o.get("buyer", {})   or {}
    trip    = o.get("trip", {})    or {}
    ff      = o.get("fulfillment", {}) or {}
    notes   = o.get("notes", []) or []

    sale   = pricing.get("quoted_price") or 0
    ctrip  = pricing.get("ctrip_price")  or 0
    status = o.get("status", "")
    profit = round(sale - ctrip * COST_RATIO, 2) if (status in PROFIT_STATUSES and sale and ctrip) else ""

    pax_n = trip.get("passenger_count", 0) or 0
    attachments = ff.get("supplier_attachments", []) if isinstance(ff, dict) else []

    note_str = "；".join(notes) if isinstance(notes, list) else str(notes or "")

    row = [
        trip.get("departure_date", ""),
        _trip_route(o),
        pax_n or "",
        buyer.get("nick", "") if isinstance(buyer, dict) else str(buyer),
        sale or "",
        ctrip or "",
        profit,
        status,
        ff.get("supplier_pnr", "") if isinstance(ff, dict) else "",
        note_str,
    ]
    return row, attachments


def _style_data_cell(cell, c_idx: int, value, status: str, attachments: list[str]) -> None:
    _, *_, align = COLUMNS[c_idx - 1]
    cell.border    = _bottom_border()
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(c_idx == COL_NOTE))

    if c_idx == COL_DEPART:
        cell.number_format = DATE_FMT
        cell.font = _font(color=C_MUTED, size=11)
    elif c_idx == COL_PNR:
        pnr_text = (value or "").strip() if isinstance(value, str) else ""
        if attachments:
            first_abs = (_ROOT / attachments[0]).resolve()
            tag = "[图]" if len(attachments) == 1 else f"[图 ×{len(attachments)}]"
            cell.value = f"{pnr_text} {tag}".strip()
            cell.hyperlink = f"file://{first_abs}"
            cell.font = _font(color=C_LINK, size=11, underline="single")
        elif pnr_text:
            cell.font = _font(color=C_MUTED, size=11)
        else:
            cell.value = ""
            cell.font = _font(color=C_MUTED, size=11)
    elif c_idx in (COL_SALE, COL_CTRIP):
        cell.number_format = MONEY_FMT
        cell.font = _font(size=12)
    elif c_idx == COL_PROFIT:
        if isinstance(value, (int, float)):
            color = C_PROFIT_POS if value >= 0 else C_PROFIT_NEG
            cell.number_format = MONEY_FMT
            cell.font = _font(color=color, size=12)
        else:
            cell.value = "—" if value == "" else value
            cell.font = _font(color=C_MUTED, size=11)
            cell.alignment = Alignment(horizontal="right", vertical="center")
    elif c_idx == COL_STATUS:
        cell.font = _font(bold=True, color=STATUS_FG.get(status, C_MUTED), size=11)
    elif c_idx == COL_NOTE:
        cell.font = _font(color=C_MUTED, size=11)
    else:
        cell.font = _font(size=12)


def write_ledger(orders: list[dict], out_path: Path) -> float:
    wb = Workbook()
    ws = wb.active
    ws.title = "闲鱼台账"
    ws.sheet_view.showGridLines = False

    rows_with_att = [order_to_row(o) for o in orders]
    n_cols = len(COLUMNS)

    ws.row_dimensions[1].height = HEADER_ROW_H
    for c_idx, (header, *_) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c_idx, value=header)
        cell.fill      = _fill(C_HEADER_BG)
        cell.font      = _font(bold=True, color=C_MUTED, size=11)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border    = _bottom_border(C_BORDER_STRONG, "medium")

    for r_idx, (row, attachments) in enumerate(rows_with_att, 2):
        ws.row_dimensions[r_idx].height = DATA_ROW_H
        status = row[COL_STATUS - 1] if len(row) >= COL_STATUS else ""
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            _style_data_cell(cell, c_idx, value, status, attachments)

    total_row = len(rows_with_att) + 2
    ws.row_dimensions[total_row].height = TOTAL_ROW_H

    rows_only = [r for r, _ in rows_with_att]
    total_sale   = sum(r[COL_SALE - 1]   for r in rows_only if isinstance(r[COL_SALE - 1],   (int, float)))
    total_cost   = sum(r[COL_CTRIP - 1]  for r in rows_only if isinstance(r[COL_CTRIP - 1],  (int, float)))
    total_profit = sum(r[COL_PROFIT - 1] for r in rows_only if isinstance(r[COL_PROFIT - 1], (int, float)))

    totals = {
        COL_DEPART: f"共 {len(rows_only)} 笔",
        COL_SALE:   total_sale,
        COL_CTRIP:  total_cost,
        COL_PROFIT: total_profit,
    }

    for c_idx in range(1, n_cols + 1):
        cell = ws.cell(row=total_row, column=c_idx, value=totals.get(c_idx, ""))
        cell.border    = _top_border()
        cell.alignment = Alignment(
            horizontal="right" if c_idx in (COL_SALE, COL_CTRIP, COL_PROFIT) else "left",
            vertical="center",
        )
        if c_idx in (COL_SALE, COL_CTRIP):
            cell.number_format = MONEY_FMT
            cell.font = _font(bold=True, size=12)
        elif c_idx == COL_PROFIT:
            color = C_PROFIT_POS if total_profit >= 0 else C_PROFIT_NEG
            cell.number_format = MONEY_FMT
            cell.font = _font(bold=True, color=color, size=12)
        else:
            cell.font = _font(bold=True, color=C_MUTED, size=11)

    for c_idx, (_, min_w, max_w, _) in enumerate(COLUMNS, 1):
        letter = get_column_letter(c_idx)
        best = float(min_w)
        for r in range(1, total_row + 1):
            v = ws.cell(row=r, column=c_idx).value
            if v is None or v == "":
                continue
            best = max(best, _cjk_width(v) * 1.2 + 2)
        ws.column_dimensions[letter].width = max(min_w, min(best, max_w))

    ws.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return total_profit


def build_ledger(out_path: Path) -> None:
    orders = list_all()
    total_profit = write_ledger(orders, out_path)
    print(f"台账文件：{out_path}")
    print(f"订单数量：{len(orders)}  累计利润：¥{total_profit:.0f}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="刷新闲鱼 Excel 台账")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    build_ledger(args.out)
